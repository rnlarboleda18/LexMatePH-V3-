import openpyxl
import re
import os

def normalize(text):
    return str(text).strip()

def extract_from_sheet(sheet):
    extracted = []
    headers = [cell.value for cell in sheet[1]]
    
    # improved column finding based on user input "case Column"
    target_idx = -1
    for i, h in enumerate(headers):
        if h and ('case' in str(h).lower() or 'title' in str(h).lower()):
            target_idx = i
            break
            
    if target_idx == -1:
        return [], f"Column not found in headers: {headers}"
        
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell_val = row[target_idx]
        if cell_val:
            val_str = str(cell_val).strip()
            # Regex for Case Number
            match = re.search(r'(G\.R\. No\.\s*[\d\s&,]+)', val_str, re.IGNORECASE)
            if match:
                # Extract and clean up the case number
                # Take the first one if multiple commas
                cn = match.group(1).split(',')[0].strip()
                extracted.append({'type': 'Number', 'value': cn, 'original': val_str})
            else:
                # Fallback to Case Name (the whole text)
                # Clean up newlines for display
                clean_name = val_str.replace('\n', ' ').strip()
                extracted.append({'type': 'Name', 'value': clean_name, 'original': val_str})
                
    return extracted, None

def main():
    file_path = r'C:\Users\rnlar\.gemini\antigravity\scratch\case_digester\Doctrinal Cases SCPh.xlsx'
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Processing: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path)
        
        all_items = []
        
        print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        print("-" * 40)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            cases, error = extract_from_sheet(sheet)
            
            if error:
                print(f"Sheet '{sheet_name}': SKIPPED - {error}")
            else:
                num_cnt = sum(1 for c in cases if c['type'] == 'Number')
                name_cnt = sum(1 for c in cases if c['type'] == 'Name')
                print(f"Sheet '{sheet_name}': {len(cases)} entries ({num_cnt} Numbers, {name_cnt} Names)")
                all_items.extend(cases)
                
        # Deduplication
        unique_values = sorted(list(set(item['value'] for item in all_items)))
        
        # Write to file
        report_path = 'data/doctrinal_cases_report.txt'
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 40 + "\n")
            f.write(f"GRAND TOTAL ENTRIES FOUND: {len(all_items)}\n")
            f.write(f"TOTAL UNIQUE ITEMS: {len(unique_values)}\n")
            f.write("=" * 40 + "\n\n")
            
            f.write("Top Unique Items:\n")
            for i, val in enumerate(unique_values, 1):
                 f.write(f"{i}. {val}\n")
                 
        print(f"Report saved to {report_path}")
             
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
