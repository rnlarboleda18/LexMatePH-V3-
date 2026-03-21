import sqlite3
import openpyxl
import os
import re

def normalize_cn(cn):
    # Simply strip whitespace for now, maybe upper case
    return str(cn).strip()

def get_sidebar_cases():
    print("Checking Sidebar Source (api/questions.db)...")
    db_path = 'api/questions.db'
    cases = []
    if not os.path.exists(db_path):
        print("  Database not found.")
        return cases
        
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        # Check table
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='doctrinal_cases'")
        if cur.fetchone():
            # Based on inspection: ['id', 'Subject', 'Case Title', 'Year', 'Key Topic', 'Doctrine / Ruling']
            cur.execute("SELECT \"Case Title\" FROM doctrinal_cases")
            
            for row in cur.fetchall():
                val = row[0]
                if val:
                     match = re.search(r'(G\.R\. No\.\s*[\d\s&,]+)', str(val))
                     if match:
                        cases.append(normalize_cn(match.group(1).split(',')[0].strip()))

        else:
            print("  Table 'doctrinal_cases' does not exist.")
        conn.close()
    except Exception as e:
        print(f"  Error reading DB: {e}")
    
    print(f"  Found {len(cases)} cases.")
    return cases

def get_excel_cases(filepath):
    print(f"Checking Excel Source ({filepath})...")
    cases = []
    
    # Handle absolute or relative paths
    if not os.path.isabs(filepath) and not filepath.startswith('data/'):
         # Fallback for old behaviour if needed, but main passes relative to root now or absolute
         pass
         
    if not os.path.exists(filepath):
        print("  File not found.")
        return cases
        
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        # Strategy 1: 'Case Title' column + Regex
        title_idx = -1
        for i, h in enumerate(headers):
             if h and 'title' in str(h).lower():
                title_idx = i
                break

        # Strategy 2: 'Case No.' column
        cn_idx = -1
        for i, h in enumerate(headers):
             if h and 'case' in str(h).lower() and 'no' in str(h).lower():
                cn_idx = i
                break
        
        # Prioritize CN column if exists, else Title Regex
        if cn_idx != -1:
             print("  Using 'Case No.' column.")
             for row in sheet.iter_rows(min_row=2, values_only=True):
                val = row[cn_idx]
                if val:
                    cases.append(normalize_cn(val))
        elif title_idx != -1:
            print("  Using 'Case Title' column regex.")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                val = row[title_idx]
                if val:
                    match = re.search(r'(G\.R\. No\.\s*[\d\s&,]+)', str(val))
                    if match:
                        cases.append(normalize_cn(match.group(1).split(',')[0].strip()))
    
    except Exception as e:
        print(f"  Error reading excel: {e}")
        
    print(f"  Found {len(cases)} cases.")
    return cases

def main():
    list1 = get_sidebar_cases()
    list2 = get_excel_cases('data/Cases_For_Manual_Digest.xlsx')
    
    # New source
    case_digester_path = r'C:\Users\rnlar\.gemini\antigravity\scratch\case_digester\Doctrinal Cases SCPh.xlsx'
    list3 = get_excel_cases(case_digester_path)
    
    all_cases = list1 + list2 + list3
    unique_cases = sorted(list(set(all_cases)))
    
    print("\n" + "="*30)
    print("CONSOLIDATED DOCTRINAL CASES REPORT")
    print("="*30)
    print(f"Sidebar Source: {len(list1)}")
    print(f"Manual Digest Excel: {len(list2)}")
    print(f"SCPh Excel (case_digester): {len(list3)}")
    print("-" * 30)
    print(f"Total Unique Cases: {len(unique_cases)}")
    print("-" * 30)
    
    for i, c in enumerate(unique_cases, 1):
        print(f"{i}. {c}")

if __name__ == "__main__":
    main()
